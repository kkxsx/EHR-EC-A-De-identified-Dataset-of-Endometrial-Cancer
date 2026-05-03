from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import os
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'
import shutil
import uvicorn
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
from tempfile import NamedTemporaryFile
from process_excel_files import process_excel_files  
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict, Any
import datetime
from openpyxl.styles import PatternFill  
from cryptography.fernet import Fernet  
import base64  
from openpyxl.styles import PatternFill
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor, as_completed
import warnings
import numpy as np
from config import dict_tables, field_config, row_rules

from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import padding
import hashlib
warnings.filterwarnings('ignore')
app = FastAPI()
# app.mount("/static", StaticFiles(directory="static"), name="static")
# templates = Jinja2Templates(directory="templates")
from magneto_aligner import align_excel_headers, apply_column_mapping

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 在生产环境中应该指定具体的域名而不是使用 "*"
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 临时保存文件的目录
UPLOAD_DIR = "uploaded_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 用于存储待修改数据的字典，key为文件名，value为待修改的数据列表
pending_updates = {}
# 分页器缓存
EXCEL_CACHE = {}

class DetectRequest(BaseModel):
    fileName: str

class PaginationRequest(BaseModel):
    fileName: str
    page: int = 1
    pageSize: int = 10
    highlightColor: Optional[str] = "FF0000"  # 默认红色（异常检测），前端可传 "FFFF00"（消歧）

class UpdateDataRequest(BaseModel):
    fileName: str
    data: List[Dict[str, Any]]
    page: int
    pageSize: int


class SaveDataRequest(BaseModel):
    fileName: str

class MaskingRequest(BaseModel):
    fileName: str
    extraColumns: Optional[List[str]] = []  # 额外指定需要脱敏的列名

class DisambiguateRequest(BaseModel):
    fileName: str
    

class MatchRequest(BaseModel):
    fileName: str
    columnMapping: Dict[str, str]  # key=源列名，value=目标列名


class CorrectionRequest(BaseModel):
    fileName: str
    actionType: str = "clear_highlights"  # 可以扩展其他操作类型
    sheetName: Optional[str] = None  # 可选，指定处理哪个工作表
    targetColor: Optional[str] = "FF0000"  # 目标颜色，默认红色

class CorrectionResponse(BaseModel):
    status: str
    message: str
    correctedFile: str
    processCount: int = -1
    statistics: Dict[str, Any]
    downloadUrl: str
class BatchCorrectionRequest(BaseModel):
    fileNames: List[str]
    actionType: str = "clear_highlights"
    targetColor: str = "FF0000"

class StatisticRequest(BaseModel):
    fileName: str

# 统计原始Excel文件信息的函数
def statistic_load_file(file_path):
    """
    统计Excel文件中的数据行数和空行数
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        dict: 包含统计数据的字典
            - total_rows: 总行数（不含表头）
            - empty_rows: 空行数
            - data_rows: 数据行数（总行数 - 空行数）
    """
    try:
        # 读取Excel文件
        df = pd.read_excel(file_path)
        
        # 总行数（不含表头）
        total_rows = len(df)
        
        # 统计空行数（所有列都为空的行）
        empty_rows = 0
        for index, row in df.iterrows():
            # 检查一行是否全部为空
            if row.isnull().all():
                empty_rows += 1
                
        # 数据行数 = 总行数 - 空行数
        data_rows = total_rows - empty_rows
        
        return {
            "total_rows": total_rows,
            "empty_rows": empty_rows,
            "data_rows": data_rows
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"统计文件信息时出错: {str(e)}")

@app.post("/api/statistic")
async def get_file_statistics(request: StatisticRequest):
    """
    获取文件统计信息接口
    """
    try:
        file_path = os.path.join(UPLOAD_DIR, request.fileName)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件未找到")
        
        # 调用统计函数
        stats = statistic_load_file(file_path)
        
        return JSONResponse(content={
            "status": "success",
            "message": "统计完成",
            "fileName": request.fileName,
            "statistics": stats
        })
        
    except Exception as e:
        print(f"文件统计失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"文件统计失败: {str(e)}")

def load_ambiguity_table(ambiguity_file_path):
    """
    加载歧义表（QuQiYi.xlsx）
    返回：dict，key=字段名（sheet名），value=dict，其中key为歧义数据，value为标准值
    """
    ambiguity_dict = {}
    
    # 读取歧义表所有sheet（每个sheet对应一个字段）
    xl_ambiguity = pd.ExcelFile(ambiguity_file_path)
    
    for sheet_name in xl_ambiguity.sheet_names:
        # 读取当前sheet的歧义数据（第一列为原始歧义数据，第二列为标准值）
        df_ambiguity = pd.read_excel(ambiguity_file_path, sheet_name=sheet_name)
        
        # 构建映射字典：歧义数据 -> 标准值
        mapping = {}
        for _, row in df_ambiguity.iterrows():
            ambiguity_value = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else None
            standard_value = row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else ambiguity_value
            
            if ambiguity_value is not None:
                # 如果标准值为空，则使用歧义值本身作为标准值
                mapping[ambiguity_value] = standard_value
        
        ambiguity_dict[sheet_name] = mapping
    
    print(f"成功加载歧义表，包含字段：{list(ambiguity_dict.keys())}")
    return ambiguity_dict

def process_single_column(args):
    """
    处理单个列的数据，用于并行计算
    返回需要标绿的单元格坐标列表 [(row, col), ...] 和需要替换的值 {(row, col): standard_value}
    """
    col_idx, col_name, series, ambiguity_mapping = args
    cells_to_highlight = []
    values_to_replace = {}
    
    if col_name in ambiguity_mapping:
        current_ambiguity_map = ambiguity_mapping[col_name]
        for row_idx, cell_value in enumerate(series, 2):  # 从第2行开始（索引为1，但Excel行号从2开始）
            cell_str = str(cell_value).strip() if pd.notna(cell_value) else ""
            if cell_str and cell_str != "nan" and cell_str in current_ambiguity_map:
                cells_to_highlight.append((row_idx, col_idx))
                # 记录需要替换的值
                values_to_replace[(row_idx, col_idx)] = current_ambiguity_map[cell_str]
                
    return cells_to_highlight, values_to_replace

def process_target_excel_multiprocessing(ambiguity_dict, target_file_path, output_file_path):
    """
    使用多进程处理目标Excel文件
    逻辑：1. 判断列是否在歧义表字段中 → 2. 并行检查数据是否在歧义表中 → 3. 符合条件标绿并替换为标准值
    """
    # 1. 用pandas读取目标Excel数据（用于快速判断数据）
    df_target = pd.read_excel(target_file_path)
    print(f"成功读取目标文件，共 {df_target.shape[0]} 行数据，{df_target.shape[1]} 列")
    
    # 2. 用openpyxl打开目标Excel（用于设置单元格颜色和替换值，需保持文件格式）
    wb = load_workbook(target_file_path)
    ws = wb.active  # 默认处理第一个sheet（如需处理所有sheet，可遍历wb.sheetnames）
    
    # 定义黄色填充样式（RGB颜色可自定义，此处为黄色）
    green_fill = PatternFill(
        start_color="FFFF00",  # 黄色RGB值
        end_color="FFFF00",
        fill_type="solid"
    )
    
    # 准备并行处理的数据
    process_args = []
    for col_idx, col_name in enumerate(df_target.columns, 1):  # col_idx：Excel列号（从1开始），col_name：列名
        # 判断当前列是否在歧义表的字段名中
        if col_name not in ambiguity_dict:
            print(f"列「{col_name}」不在歧义表字段中，跳过处理")
            continue
        
        print(f"准备处理列「{col_name}」...")
        # 添加到处理参数列表
        process_args.append((col_idx, col_name, df_target[col_name], ambiguity_dict))
    
    # 使用多进程并行处理所有列
    cells_to_highlight = []
    all_values_to_replace = {}
    cpu_count = min(mp.cpu_count(), len(process_args))  # 根据任务数确定使用的进程数
    
    if cpu_count > 1 and len(process_args) > 1:
        print(f"使用 {cpu_count} 个进程并行处理 {len(process_args)} 列...")
        with ProcessPoolExecutor(max_workers=cpu_count) as executor:
            futures = [executor.submit(process_single_column, args) for args in process_args]
            for future in as_completed(futures):
                highlighted, to_replace = future.result()
                cells_to_highlight.extend(highlighted)
                all_values_to_replace.update(to_replace)
    else:
        # 如果只有单列或单核CPU，直接顺序处理
        print("顺序处理所有列...")
        for args in process_args:
            highlighted, to_replace = process_single_column(args)
            cells_to_highlight.extend(highlighted)
            all_values_to_replace.update(to_replace)
    
    # 应用高亮样式和值替换
    print(f"正在应用高亮样式到 {len(cells_to_highlight)} 个单元格...")
    print(f"正在替换 {len(all_values_to_replace)} 个单元格的值...")
    
    # 替换值
    for (row_idx, col_idx), standard_value in all_values_to_replace.items():
        ws.cell(row=row_idx, column=col_idx).value = standard_value
    
    # 应用高亮样式
    for row_idx, col_idx in cells_to_highlight:
        ws.cell(row=row_idx, column=col_idx).fill = green_fill
    
    # 4. 保存处理后的文件（不覆盖原文件，输出到新路径）
    wb.save(output_file_path)
    print(f"\n处理完成！结果已保存至：{output_file_path}")

# 去歧义处理接口
@app.post("/api/disambiguate")
async def disambiguate(request: DisambiguateRequest):
    try:
        # 待处理文件名（已上传到 uploads）
        input_filename = request.fileName
        input_path = os.path.join(UPLOAD_DIR, input_filename)

        if not os.path.exists(input_path):
            raise HTTPException(status_code=404, detail=f"文件未找到：{input_filename}")

        # 歧义表文件（固定放 uploads 中）
        ambiguity_file = os.path.join(os.path.dirname(__file__), "QuQiYi.xlsx")
        if not os.path.exists(ambiguity_file):
            raise HTTPException(status_code=404, detail="未找到歧义表文件 QuQiYi.xlsx，请确保文件存在")

        # 生成输出文件
        base_name, ext = os.path.splitext(input_filename)
        output_filename = f"{base_name}_quqiyi{ext}"
        output_path = os.path.join(UPLOAD_DIR, output_filename)

        # 加载歧义表
        ambiguity_dict = load_ambiguity_table(ambiguity_file)
        
        # 调用消歧脚本逻辑
        process_target_excel_multiprocessing(
            ambiguity_dict=ambiguity_dict,
            target_file_path=input_path,
            output_file_path=output_path
        )
        
        # 统计被标记的单元格数量（重新打开文件进行计数）
        wb = load_workbook(output_path)
        ws = wb.active
        cells_highlighted = 0
        green_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill.fgColor.rgb == green_fill.start_color.index:
                    cells_highlighted += 1

        # 返回值（按你要求）
        response_content = {
            "status": "success",
            "message": f"成功标记 {cells_highlighted} 个歧义单元格",
            "outputFile": output_filename,
            "downloadUrl": f"/api/download/{output_filename}",
            "cellsHighlighted": cells_highlighted,
            "processCount": cells_highlighted
        }

        return JSONResponse(content=response_content)

    except Exception as e:
        print(f"去歧义处理失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"去歧义处理失败: {str(e)}")



#数据脱敏（不加盐）
def mask_sensitive_data(input_xlsx, output_xlsx, cols, key_file):
    """对Excel文件中的指定列进行AES加密脱敏（确定性加密）"""
    ENC_PREFIX = "ENC::"
    
    # 生成或读取密钥
    def get_or_create_key(key_file):
        if os.path.exists(key_file):
            with open(key_file, "rb") as f:
                key = f.read()
        else:
            # 生成32字节密钥（AES-256）
            key = os.urandom(32)
            with open(key_file, "wb") as f:
                f.write(key)
        return key

    # 确定性加密函数
    def encrypt_value(key, plaintext):
        if pd.isna(plaintext):
            return plaintext
        plaintext = str(plaintext)
        
        # 使用明文的哈希值作为固定IV（确保相同明文使用相同IV）
        iv = hashlib.sha256(plaintext.encode('utf-8')).digest()[:16]
        
        # 创建加密器
        cipher = Cipher(
            algorithms.AES(key),
            modes.CBC(iv),
            backend=default_backend()
        )
        encryptor = cipher.encryptor()
        
        # 填充明文到16字节的倍数
        padder = padding.PKCS7(128).padder()
        padded_data = padder.update(plaintext.encode('utf-8')) + padder.finalize()
        
        # 加密
        ciphertext = encryptor.update(padded_data) + encryptor.finalize()
        
        return ENC_PREFIX + base64.b64encode(ciphertext).decode("utf-8")

    # 读取数据
    df = pd.read_excel(input_xlsx, dtype=str, keep_default_na=False)
    key = get_or_create_key(key_file)

    # 对指定列进行加密
    for col in cols:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: encrypt_value(key, x))
        else:
            print(f"⚠️ 列 {col} 不存在，跳过。")

    # 保存加密后的文件
    df.to_excel(output_xlsx, index=False, engine="openpyxl")
    print(f"✅ 已输出脱敏文件：{output_xlsx}")
    print(f"🔑 密钥文件：{key_file}")


def mask_sensitive_data_blank(input_xlsx, output_xlsx, cols, key_file):
    """按 mask_sensitive_data 相同入参，直接将指定列内容置空。"""
    # 为保持调用签名兼容，保留 key_file 参数但不使用。
    _ = key_file

    # 读取数据并保持原始字符串形态，避免被自动转成 NaN。
    df = pd.read_excel(input_xlsx, dtype=str, keep_default_na=False)

    # 对指定列执行直接置空。
    for col in cols:
        if col in df.columns:
            df[col] = ""
        else:
            print(f"⚠️ 列 {col} 不存在，跳过。")

    # 保存置空后的文件。
    df.to_excel(output_xlsx, index=False, engine="openpyxl")
    print(f"✅ 已输出置空脱敏文件：{output_xlsx}")

# 自动识别敏感列
def identify_sensitive_columns(columns):
    """根据列名特征自动识别敏感列"""
    sensitive_columns = []
    
    # 关键词匹配模式
    sensitive_keywords = ['phone', 'address', 'contact']
    
    for col in columns:
        col_lower = col.lower()
        # 检查是否包含敏感关键词
        if any(keyword in col_lower for keyword in sensitive_keywords):
            sensitive_columns.append(col)
            
    return sensitive_columns

# 数据脱敏接口
@app.post("/api/mask")
async def mask_data(request: MaskingRequest):
    """对上传的Excel文件进行数据脱敏"""
    try:
        file_path = os.path.join(UPLOAD_DIR, request.fileName)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件未找到，请先上传文件")
        
        # 检查文件扩展名
        file_extension = request.fileName.split('.')[-1]
        if file_extension not in ['xlsx', 'xls']:
            raise HTTPException(status_code=400, detail="只支持处理 Excel 文件（.xlsx, .xls）")
        
        print(f"开始对文件进行数据脱敏: {request.fileName}")
        
        # 读取文件以获取列名
        df = pd.read_excel(file_path, nrows=0)  # 只读取表头
        all_columns = list(df.columns)
        
        # 预设的需要脱敏的列（根据用户提供的信息）
        preset_sensitive_columns = [
            'SICK_NAME', 'ID_NUMBER', 'CONTACT_ADDRESS', 'CONTACT_NUMBER', 
            'CONTACTS', 'DOCTOR', 'APPLY_DOCTOR', 'APPLY_OPERATOR', 
            'EXAM_OPERATOR', 'REPORT_OPERATOR', 'AUDIT_OPERATOR', 
            'PRINT_OPERATOR', 'SUBSEQUENT_VISIT_OPERATOR', 'SLICE_OPERATOR',
            'DIAG_DOCTOR', 'OPERATOR', 'INSPECT_APPLY_DOCTOR'
        ]
        
        # 自动识别敏感列
        auto_sensitive_columns = identify_sensitive_columns(all_columns)
        
        # 合并所有需要脱敏的列
        sensitive_columns = []
        
        # 添加预设的敏感列（如果存在于文件中）
        for col in preset_sensitive_columns:
            if col in all_columns and col not in sensitive_columns:
                sensitive_columns.append(col)
        
        # 添加自动识别的敏感列
        for col in auto_sensitive_columns:
            if col not in sensitive_columns:
                sensitive_columns.append(col)
        
        # 添加前端指定的额外列
        for col in request.extraColumns:
            if col in all_columns and col not in sensitive_columns:
                sensitive_columns.append(col)
        
        if not sensitive_columns:
            raise HTTPException(status_code=400, detail="未找到需要脱敏的列")
        
        print(f"需要脱敏的列: {sensitive_columns}")
        
        # 构造输出文件名和密钥文件名
        base_name = os.path.splitext(request.fileName)[0]
        masked_file = os.path.join(UPLOAD_DIR, f"{base_name}_masked.xlsx")
        key_file = os.path.join(UPLOAD_DIR, f"{base_name}_key.bin")
        
        # # 执行数据脱敏
        # mask_sensitive_data(file_path, masked_file, sensitive_columns, key_file)
        # 执行数据脱敏（直接置空版本）
        mask_sensitive_data_blank(file_path, masked_file, sensitive_columns, key_file)
        # 检查结果文件是否存在
        if not os.path.exists(masked_file):
            raise HTTPException(status_code=500, detail="脱敏后的文件未生成")
            
        # if not os.path.exists(key_file):
        #     raise HTTPException(status_code=500, detail="密钥文件未生成")
        
        # 返回结果
        masked_filename = os.path.basename(masked_file)
        key_filename = os.path.basename(key_file)
        
        return JSONResponse(content={
            "status": "success",
            "message": f"成功对{len(sensitive_columns)}列进行脱敏处理",
            "maskedFile": masked_filename,
            "keyFile": key_filename,
            "sensitiveColumns": sensitive_columns,
            "processedCount": len(sensitive_columns),
            "downloadUrls": {
                "maskedFile": f"/api/download/{masked_filename}",
                "keyFile": f"/api/download/{key_filename}"
            }
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"数据脱敏失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"数据脱敏失败: {str(e)}")
#前端上传文件
@app.post("/api/upload")
async def upload_file(files: UploadFile = File(...)):
    print(f"收到文件: {files.filename}, 大小: {files.size}")
    try:
        file_name = files.filename
        file_extension = file_name.split('.')[-1].lower()

        if file_extension not in ['xlsx', 'xls']:
            raise HTTPException(status_code=400, detail="只支持上传 Excel 文件（.xlsx, .xls）")

        temp_file_path = os.path.join(UPLOAD_DIR, file_name)
        print(f"临时文件保存路径: {temp_file_path}")

        # 保存上传文件
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(files.file, buffer)

        # === 新增：列名智能对齐层（不会破坏原业务逻辑）===
        alignment_meta = None
        try:
            # ⚠️ openpyxl 不能直接处理 .xls（老格式）
            if file_extension == "xlsx":
                alignment_meta = align_excel_headers(
                    temp_file_path,
                    exception_map_path=os.path.join(os.path.dirname(__file__), "exception_map.xlsx"),
                    model_name="BAAI/bge-base-zh-v1.5",
                    threshold_auto=0.75,      # 自动改名阈值（保守，医疗推荐）
                    threshold_suggest=0.65,   # 建议阈值（只建议不写回）
                    top_k=5,
                    sidecar_dir=UPLOAD_DIR,
                    sheet_index=0,
                    write_back=False,         # 仅计算匹配，不自动写回
                )

                # 不再清除缓存，因为文件未被修改
            else:
                alignment_meta = {
                    "status": "skipped",
                    "reason": "xls 格式不支持直接写回表头（openpyxl 不支持），请转为 xlsx 后再上传以启用智能对齐。",
                    "mapping": []
                }

        except Exception as e:
            # ✅ 不影响上传成功：对齐失败也放行
            print(f"列名对齐失败（不影响上传）: {str(e)}")
            alignment_meta = {
                "status": "failed",
                "error": str(e),
                "mapping": []
            }

        return JSONResponse(content={
            "filename": file_name,
            "message": "文件上传成功",
            "status": "success",
            "alignment": alignment_meta,   # ✅ 给前端展示对齐结果
        })

    except Exception as e:
        print(f"文件上传失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")

# 应用字段映射接口
@app.post("/api/match")
async def apply_field_mapping(request: MatchRequest):
    """
    根据前端用户确认的字段映射，物理修改 Excel 文件的列名
    
    - **fileName**: 要处理的文件名（需已上传）
    - **columnMapping**: 列名映射字典，key=源列名，value=目标列名
    
    返回修改后的完整列名列表
    """
    try:
        file_path = os.path.join(UPLOAD_DIR, request.fileName)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件未找到，请先上传文件")
        
        # 检查文件扩展名
        file_extension = request.fileName.split('.')[-1].lower()
        if file_extension not in ['xlsx', 'xls']:
            raise HTTPException(status_code=400, detail="只支持处理 Excel 文件（.xlsx, .xls）")
        
        if file_extension == 'xls':
            raise HTTPException(status_code=400, detail="暂不支持 .xls 格式的字段映射，请转换为 .xlsx 格式")
        
        print(f"开始应用字段映射: {request.fileName}")
        print(f"映射关系: {request.columnMapping}")
        
        # 调用 apply_column_mapping 函数
        result = apply_column_mapping(
            file_path=file_path,
            column_mapping=request.columnMapping,
            sheet_index=0
        )
        
        if not result["success"]:
            raise HTTPException(
                status_code=400, 
                detail=result["message"] if "message" in result else "字段映射失败"
            )
        
        # 修改成功后，清除缓存（保证 paging 接口读取到新列名）
        EXCEL_CACHE.pop(file_path, None)
        
        return JSONResponse(content={
            "success": True,
            "message": result["message"],
            "newColumns": result["newColumns"],
            "mappedCount": result["mappedCount"],
            "errors": result.get("errors", [])
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"应用字段映射失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"应用字段映射失败: {str(e)}")

def get_cached_df(file_path):
    # 获取文件修改时间
    mtime = os.path.getmtime(file_path)

    # 缓存命中：文件存在且未修改
    if file_path in EXCEL_CACHE:
        cached = EXCEL_CACHE[file_path]
        if cached["mtime"] == mtime:
            return cached["df"]

    # 缓存未命中或文件已更新：重新读取
    df = pd.read_excel(file_path)

    EXCEL_CACHE[file_path] = {
        "mtime": mtime,
        "df": df
    }
    return df


@app.post("/api/paging")
async def get_excel_data(request: PaginationRequest):
    try:
        file_path = os.path.join(UPLOAD_DIR, request.fileName)

        if not os.path.exists(file_path):
            print(f"文件未找到: {file_path}")
            raise HTTPException(status_code=404, detail="文件未找到")

        # ============================================================
        # 1. DataFrame 走缓存（保持不动）
        # ============================================================
        df = get_cached_df(file_path)

        # ============================================================
        # 2. 分页计算（保持不动）
        # ============================================================
        total_rows = len(df)
        total_pages = (total_rows + request.pageSize - 1) // request.pageSize
        columns = list(df.columns)

        start_idx = (request.page - 1) * request.pageSize
        end_idx = min(start_idx + request.pageSize, total_rows)
        page_data = df.iloc[start_idx:end_idx].copy()

        # ============================================================
        # 3. 读取 Excel 样式（仅用于高亮）
        # ============================================================
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active

        highlighted_cells = []
        
        # 使用请求中的颜色参数（支持异常检测红色和消歧黄色）
        target_color = request.highlightColor.upper()

        # Excel 中：第 1 行是表头，数据从第 2 行开始
        excel_start_row = start_idx + 2
        excel_end_row = end_idx + 2

        for row_idx in range(excel_start_row, excel_end_row):
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if (
                    cell.fill
                    and cell.fill.patternType == "solid"
                    and cell.fill.start_color
                ):
                    color_rgb = cell.fill.start_color.rgb
                    # 使用参数化的颜色检测
                    if color_rgb and str(color_rgb).upper().endswith(target_color):
                        highlighted_cells.append({
                            "row": row_idx - 2,      # DataFrame 行索引（0-based）
                            "column": col_name,      # 列名
                            "address": cell.coordinate
                        })

        # ============================================================
        # 4. 统一安全转换函数（保持不动）
        # ============================================================
        def safe_convert(v):
            if pd.isna(v):
                return ""

            # 日期时间类型
            if isinstance(v, (pd.Timestamp, np.datetime64)):
                try:
                    return pd.to_datetime(v).isoformat()
                except:
                    return ""

            if isinstance(v, (datetime.date, datetime.time, datetime.datetime)):
                try:
                    return v.isoformat()
                except:
                    return ""

            # 数值类型
            if isinstance(v, np.integer):
                return int(v)

            if isinstance(v, np.floating):
                if np.isnan(v):
                    return ""
                return float(v)

            # 布尔类型
            if isinstance(v, np.bool_):
                return bool(v)

            return v

        # ============================================================
        # 5. 对分页数据做统一转换（保持不动）
        # ============================================================
        for col in page_data.columns:
            page_data[col] = page_data[col].apply(safe_convert)

        page_data = page_data.replace({pd.NaT: "", np.nan: ""}).fillna("")

        data = page_data.to_dict(orient="records")

        # ============================================================
        # 6. 返回结果（新增 highlightedCells）
        # ============================================================
        return JSONResponse(content={
            "filename": request.fileName,
            "columns": columns,
            "data": data,
            "currentPage": request.page,
            "totalPages": total_pages,
            "totalRows": total_rows,
            "pageSize": request.pageSize,
            "highlightedCells": highlighted_cells
        })

    except Exception as e:
        print(f"获取Excel数据失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取Excel数据失败: {str(e)}")




# 更新Excel数据的接口
@app.post("/api/update")
async def update_excel_data(request: UpdateDataRequest):
    try:
        file_path = os.path.join(UPLOAD_DIR, request.fileName)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件未找到")
        
        # 将待修改数据暂存到内存中，而不是立即写入文件
        if request.fileName not in pending_updates:
            pending_updates[request.fileName] = []
        
        # 添加新的待修改数据
        for row_data in request.data:
            pending_updates[request.fileName].append({
                "data": row_data,
                "page": request.page,
                "pageSize": request.pageSize
            })
        
        return JSONResponse(content={
            "message": "数据已暂存，等待保存",
            "status": "success"
        })
        
    except Exception as e:
        print(f"暂存Excel数据失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"暂存Excel数据失败: {str(e)}")

# 新增保存数据接口
@app.post("/api/save")
async def save_excel_data(request: SaveDataRequest):
    try:
        file_path = os.path.join(UPLOAD_DIR, request.fileName)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件未找到")
        
        if request.fileName not in pending_updates or not pending_updates[request.fileName]:
            return JSONResponse(content={
                "message": "没有待保存的数据",
                "status": "success"
            })
        
        # 读取Excel文件
        df = pd.read_excel(file_path)
        
        # 应用所有暂存的修改
        for update_item in pending_updates[request.fileName]:
            row_data = update_item["data"]
            page = update_item["page"]
            pageSize = update_item["pageSize"]
            
            # 计算当前页在文件中的起始和结束索引
            start_idx = (page - 1) * pageSize
            
            # 从数据中获取实际的行索引
            if '__rowIndex' in row_data:
                row_index = row_data['__rowIndex']
                # 移除辅助字段
                row_data_without_index = {k: v for k, v in row_data.items() if k != '__rowIndex'}
                
                # 确保行索引在有效范围内
                if 0 <= row_index < len(df):
                    for col, value in row_data_without_index.items():
                        if col in df.columns:
                            # 处理空字符串为NaN
                            if value == "":
                                df.at[row_index, col] = pd.NA
                            else:
                                df.at[row_index, col] = value
            else:
                # 如果没有提供行索引，使用基于页面的计算
                # 这种情况是为了兼容旧的调用方式
                relative_index = 0
                for i, item in enumerate(pending_updates[request.fileName]):
                    if item["data"] is row_data:
                        relative_index = i
                        break
                row_index = start_idx + relative_index
                
                if 0 <= row_index < len(df):
                    for col, value in row_data.items():
                        if col in df.columns:
                            # 处理空字符串为NaN
                            if value == "":
                                df.at[row_index, col] = pd.NA
                            else:
                                df.at[row_index, col] = value
        
        # 保存更新后的数据到原文件
        df.to_excel(file_path, index=False)
        
        # 清除已保存的暂存数据
        pending_updates.pop(request.fileName, None)
        
        return JSONResponse(content={
            "message": "数据保存成功",
            "status": "success"
        })
        
    except Exception as e:
        print(f"保存Excel数据失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"保存Excel数据失败: {str(e)}")


 
# 异常检测接口 - 
@app.post("/api/detect")
async def detect(request: DetectRequest):
    print(f"收到处理请求，文件名: {request.fileName}")
    try:
        file_path = os.path.join(UPLOAD_DIR, request.fileName)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件未找到，请先上传文件")
        from main import process_excel_files as mask
        # 调用 process_excel_files 进行处理
        result_files = mask(
            data_files=[file_path],
            exception_code_file="exception_map.xlsx",
            dict_tables=dict_tables,
            field_config=field_config,
            row_rules=row_rules
        )
        return JSONResponse(content={
                "status": "success", 
                "message": "异常检测完成",
                "resultFilename": os.path.basename(result_files[0]["highlight_file"])
                    
        })
        # if result_files and len(result_files) > 0:
        #     # 原始生成的路径（可能是 static/xxx.xlsx）
        #     original_result_path = result_files[0]["highlight_file"]
            
        #     if os.path.exists(original_result_path):
        #         # 获取文件名
        #         result_filename = os.path.basename(original_result_path)
                
        #         # 【关键步骤】将结果文件移动到 UPLOAD_DIR，以便 download 接口能找到它
        #         target_path = os.path.join(UPLOAD_DIR, result_filename)
                
        #         # 如果文件不在目标目录，则移动或复制过去
        #         if os.path.abspath(original_result_path) != os.path.abspath(target_path):
        #             shutil.copy2(original_result_path, target_path)
                
        #         print(f"处理完成，新文件名为: {result_filename}")
                
        #         # 【修改点】返回 JSON，包含新的文件名
        #         return JSONResponse(content={
        #             "status": "success", 
        #             "message": "异常检测完成",
        #             "resultFilename": result_filename 
        #         })
        #     else:
        #         raise HTTPException(status_code=500, detail="处理后的文件生成失败")
        # else:
        #     raise HTTPException(status_code=500, detail="文件处理未产生结果")
            
    except Exception as e:
        print(f"文件处理失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"文件处理失败: {str(e)}")



def clear_highlighted_content(file_path: str, sheet_name: Optional[str] = None, target_color: str = "FF0000") -> Dict[
    str, Any]:
    """
    读取 Excel 文件，查找特定颜色高亮背景的单元格，并清空其内容。

    Args:
        file_path: Excel 文件路径
        sheet_name: 可选，指定处理的工作表名称，如果为None则处理所有工作表
        target_color: 目标颜色代码，默认红色 "FF0000"

    Returns:
        包含统计信息的字典
    """
    print(f"正在处理文件: {file_path}")

    try:
        # 加载工作簿 (data_only=False 以保留公式和样式)
        wb = load_workbook(file_path, data_only=False)
    except FileNotFoundError:
        raise Exception(f"找不到文件 {file_path}")

    statistics = {
        "total_cells_cleared": 0,
        "sheets_processed": [],
        "details": {}
    }

    # 确定要处理的工作表
    sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames

    # 遍历指定的工作表
    for sheet_name in sheets_to_process:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        print(f"  - 正在扫描工作表: {sheet_name}")

        count = 0
        cleared_cells = []

        # 遍历工作表中所有已使用的单元格
        for row in ws.iter_rows():
            for cell in row:
                # 获取填充样式
                fill = cell.fill
                if fill and fill.patternType == 'solid' and fill.start_color:
                    color_rgb = fill.start_color.rgb

                    # 检查颜色是否匹配目标颜色
                    if color_rgb and str(color_rgb).endswith(target_color.upper()):
                        # 记录被清空的单元格信息
                        cell_info = {
                            "address": cell.coordinate,
                            "column": get_column_letter(cell.column),
                            "row": cell.row,
                            "original_value": str(cell.value) if cell.value else "",
                            "color": str(color_rgb)
                        }
                        cleared_cells.append(cell_info)

                        # 清空内容
                        cell.value = None
                        count += 1

        print(f"  - 已清空 {count} 个高亮单元格的内容")

        # 记录统计信息
        statistics["sheets_processed"].append(sheet_name)
        statistics["details"][sheet_name] = {
            "cells_cleared": count,
            "cleared_cells": cleared_cells
        }
        statistics["total_cells_cleared"] += count

    # 生成输出文件名
    base_name = os.path.splitext(file_path)[0]
    output_file = f"{base_name}_corrected.xlsx"

    # 保存文件
    wb.save(output_file)
    print(f"✅ 处理完成，结果已保存至: {output_file}\n")

    return {
        "output_file": output_file,
        "statistics": statistics
    }

#异常处理接口
@app.post("/api/correct", response_model=CorrectionResponse)
async def correct_data(request: CorrectionRequest):
    """
    异常处理接口 - 清理高亮单元格内容

    - **fileName**: 要处理的文件名（需已上传）
    - **actionType**: 操作类型，目前支持 "clear_highlights"
    - **sheetName**: 可选，指定要处理的工作表名称
    - **targetColor**: 可选，目标颜色代码，默认红色 "FF0000"

    返回处理结果文件下载链接和统计信息
    """
    try:
        # 构建文件路径
        file_path = os.path.join(UPLOAD_DIR, request.fileName)

        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件未找到，请先上传文件")

        # 检查文件扩展名
        file_extension = request.fileName.split('.')[-1]
        if file_extension not in ['xlsx', 'xls']:
            raise HTTPException(status_code=400, detail="只支持处理 Excel 文件（.xlsx, .xls）")

        print(f"开始执行异常处理: {request.fileName}")
        print(f"操作类型: {request.actionType}, 目标颜色: {request.targetColor}")

        # 根据操作类型执行不同的处理
        if request.actionType == "clear_highlights":
            # 调用清理高亮内容函数
            result = clear_highlighted_content(
                file_path=file_path,
                sheet_name=request.sheetName,
                target_color=request.targetColor
            )

            output_file = result["output_file"]
            statistics = result["statistics"]

            # 检查输出文件是否存在
            if not os.path.exists(output_file):
                raise HTTPException(status_code=500, detail="处理后的文件未生成")

            # 获取相对路径用于下载
            output_filename = os.path.basename(output_file)

            # 如果需要，可以将处理后的文件移动到上传目录
            dest_path = os.path.join(UPLOAD_DIR, output_filename)
            if output_file != dest_path:
                shutil.copy2(output_file, dest_path)
                os.remove(output_file)  # 删除临时文件
                output_file = dest_path

            return CorrectionResponse(
                status="success",
                message=f"成功清理 {statistics['total_cells_cleared']} 个高亮单元格",
                correctedFile=output_filename,
                processCount=statistics['total_cells_cleared'],
                statistics=statistics,
                
                downloadUrl=f"/api/download/{output_filename}"
            )
        else:
            raise HTTPException(status_code=400, detail=f"不支持的操作类型: {request.actionType}")

    except HTTPException:
        raise
    except Exception as e:
        print(f"异常处理失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"异常处理失败: {str(e)}")


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """
    文件下载接口
    """
    try:
        file_path = os.path.join(UPLOAD_DIR, filename)

        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件不存在")

        return FileResponse(
            file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")




#启动逻辑
if __name__ == "__main__":
    # host="0.0.0.0" 允许外部访问，port 是后端端口（和前端代理配置一致）
    uvicorn.run(app, host="0.0.0.0", port=3000)