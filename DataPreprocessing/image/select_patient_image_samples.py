#!/usr/bin/env python3
import argparse
import os
import random
import shutil
from pathlib import Path
from typing import Dict, Iterable, List

# Common image-like extensions. Empty suffix is also accepted (many DICOM files have no extension).
IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".svs", ".ndpi",
    ".dcm", ".dicom", ".webp", ".gif", ".heic", ".nii", ".nii.gz",
}

# Files that are very likely metadata/non-image.
NON_IMAGE_EXTENSIONS = {
    ".txt", ".csv", ".json", ".xml", ".yaml", ".yml", ".md", ".log",
    ".zip", ".rar", ".7z", ".tar", ".gz", ".pdf",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Scan By_Sickid directory, classify patients into three groups "
            "(only ultrasound, only pathology, both), sample diverse cases by image count, "
            "and export to an xlsx with three sheets."
        )
    )
    parser.add_argument(
        "--root",
        type=Path,
        required=True,
        help="Path to By_Sickid root, e.g. /media/.../data_med/By_Sickid",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("patient_image_samples.xlsx"),
        help="Output xlsx path",
    )
    parser.add_argument(
        "--n-per-type",
        type=int,
        default=12,
        help="How many patients to sample for each category",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Random seed for tie-breaking/random fill",
    )
    parser.add_argument(
        "--max-example-paths",
        type=int,
        default=5,
        help="Max number of example file paths shown per modality in each row",
    )
    parser.add_argument(
        "--unique-counts",
        action="store_true",
        help=(
            "Pick one patient per distinct image-count value (ascending). "
            "Example with --n-per-type 30 and --count-start 1: use 30 different count values."
        ),
    )
    parser.add_argument(
        "--count-start",
        type=int,
        default=1,
        help="Start count value for --unique-counts mode (default: 1).",
    )
    parser.add_argument(
        "--export-sample-dir",
        type=Path,
        default=None,
        help=(
            "If set, copy selected image files into this directory with layout: "
            "<sheet>/<count>/<sickid>/<original_rel_path>."
        ),
    )
    return parser.parse_args()


def _is_image_like(path: Path) -> bool:
    suffix = path.suffix.lower()
    if suffix in NON_IMAGE_EXTENSIONS:
        return False
    if suffix in IMAGE_EXTENSIONS:
        return True
    # For extensionless files (common in medical imaging), keep them.
    if suffix == "":
        return True
    # Unknown suffix: conservatively keep it if file is under modality folder.
    return True


def _walk_image_files(base_dir: Path) -> List[Path]:
    files: List[Path] = []
    if not base_dir.exists() or not base_dir.is_dir():
        return files

    # followlinks=True handles datasets where modality folders are symbolic links.
    for root, _, filenames in os.walk(str(base_dir), followlinks=True):
        root_path = Path(root)
        for name in filenames:
            fp = root_path / name
            if _is_image_like(fp):
                files.append(fp)
    return files


def _is_pathology_dirname(name: str) -> bool:
    lower = name.lower()
    keywords = ["pathology", "patho", "病理"]
    return any(k in lower for k in keywords)


def _collect_files_by_modality(patient_dir: Path) -> Dict[str, List[Path]]:
    # User-confirmed strong rule: if B-ultrasound data exists, it is under <sickid>/dicom/.
    b_files = _walk_image_files(patient_dir / "dicom")

    # Pathology folders may vary; collect from directory names that indicate pathology.
    p_files: List[Path] = []
    for root, dirnames, _ in os.walk(str(patient_dir), followlinks=True):
        root_path = Path(root)
        for dirname in dirnames:
            if _is_pathology_dirname(dirname):
                p_files.extend(_walk_image_files(root_path / dirname))

    return {
        "b_files": sorted(set(b_files)),
        "pathology_files": sorted(set(p_files)),
    }


def _format_example_paths(paths: Iterable[Path], max_n: int, patient_dir: Path) -> str:
    items = list(paths)
    if not items:
        return ""
    shown = items[:max_n]
    rels = [str(p.relative_to(patient_dir)) for p in shown]
    if len(items) > max_n:
        rels.append(f"... (+{len(items) - max_n} more)")
    return "\n".join(rels)


def _safe_rel_path(src: Path, patient_dir: Path) -> Path:
    try:
        return src.relative_to(patient_dir)
    except ValueError:
        return Path(src.name)


def _scan_patients(root: Path, max_example_paths: int) -> List[Dict]:
    rows: List[Dict] = []

    patient_dirs = sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name)
    for patient_dir in patient_dirs:
        sickid = patient_dir.name
        files = _collect_files_by_modality(patient_dir)
        b_files = files["b_files"]
        p_files = files["pathology_files"]

        row = {
            "sickid": sickid,
            "has_b_ultrasound": bool(b_files),
            "has_pathology": bool(p_files),
            "b_image_count": len(b_files),
            "pathology_image_count": len(p_files),
            "total_image_count": len(b_files) + len(p_files),
            "b_example_paths": _format_example_paths(b_files, max_example_paths, patient_dir),
            "pathology_example_paths": _format_example_paths(p_files, max_example_paths, patient_dir),
            "patient_dir": str(patient_dir),
            "b_files": b_files,
            "pathology_files": p_files,
        }
        rows.append(row)

    return rows


def _pick_diverse(
    records: List[Dict],
    n: int,
    count_key: str,
    seed: int,
    unique_counts: bool = False,
    count_start: int = 1,
) -> List[Dict]:
    if n <= 0 or not records:
        return []

    if unique_counts:
        count_to_rows: Dict[int, List[Dict]] = {}
        for row in records:
            cnt = int(row[count_key])
            count_to_rows.setdefault(cnt, []).append(row)

        selected: List[Dict] = []
        used_counts = set()

        # First pass: prefer counts >= count_start, ascending; skip missing counts naturally.
        for cnt in sorted(c for c in count_to_rows.keys() if c >= count_start):
            candidate = sorted(count_to_rows[cnt], key=lambda x: x["sickid"])[0]
            selected.append(candidate)
            used_counts.add(cnt)
            if len(selected) >= n:
                return sorted(selected, key=lambda x: (x[count_key], x["sickid"]))

        # Second pass: if still short, use smaller counts (still non-repeating by count).
        for cnt in sorted(c for c in count_to_rows.keys() if c < count_start):
            if cnt in used_counts:
                continue
            candidate = sorted(count_to_rows[cnt], key=lambda x: x["sickid"])[0]
            selected.append(candidate)
            used_counts.add(cnt)
            if len(selected) >= n:
                return sorted(selected, key=lambda x: (x[count_key], x["sickid"]))

        # Not enough unique count values in this category.
        return sorted(selected, key=lambda x: (x[count_key], x["sickid"]))

    if len(records) <= n:
        return sorted(records, key=lambda x: (x[count_key], x["sickid"]))

    ordered = sorted(records, key=lambda x: (x[count_key], x["sickid"]))

    # Evenly pick positions from low to high count, so we naturally cover sparse/medium/dense cases.
    if n == 1:
        indices = [len(ordered) // 2]
    else:
        indices = [round(i * (len(ordered) - 1) / (n - 1)) for i in range(n)]

    chosen: List[Dict] = []
    seen = set()
    for idx in indices:
        sid = ordered[idx]["sickid"]
        if sid not in seen:
            chosen.append(ordered[idx])
            seen.add(sid)

    if len(chosen) < n:
        rng = random.Random(seed)
        remaining = [r for r in ordered if r["sickid"] not in seen]
        rng.shuffle(remaining)
        chosen.extend(remaining[: (n - len(chosen))])

    return sorted(chosen, key=lambda x: (x[count_key], x["sickid"]))


def _fit_sheet_style(workbook, font_cls) -> None:
    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"

        # Header style
        for cell in ws[1]:
            cell.font = font_cls(name="Arial", size=11, bold=True)

        # Body font + adaptive column width
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font_cls(name="Arial", size=11)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value.split("\n")[0]))
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def _export_sample_files(
    rows: List[Dict],
    sheet_name: str,
    count_key: str,
    export_root: Path,
    include_b: bool,
    include_pathology: bool,
) -> int:
    copied = 0
    for row in rows:
        sickid = str(row["sickid"])
        count_value = int(row[count_key])
        patient_dir = Path(str(row["patient_dir"]))

        src_files: List[Path] = []
        if include_b:
            src_files.extend(row.get("b_files", []))
        if include_pathology:
            src_files.extend(row.get("pathology_files", []))

        for src in sorted(set(src_files)):
            if not src.exists() or not src.is_file():
                continue
            rel = _safe_rel_path(src, patient_dir)
            dst = export_root / sheet_name / str(count_value) / sickid / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(str(src), str(dst))
            copied += 1
    return copied


def main() -> None:
    args = parse_args()

    root = args.root.expanduser().resolve()
    if not root.exists() or not root.is_dir():
        raise FileNotFoundError(f"Root path not found or not a directory: {root}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ModuleNotFoundError(
            "openpyxl is required to export xlsx. Please install it with: pip3 install openpyxl"
        ) from exc

    all_rows = _scan_patients(root, args.max_example_paths)

    only_b = [r for r in all_rows if r["has_b_ultrasound"] and not r["has_pathology"]]
    only_pathology = [r for r in all_rows if r["has_pathology"] and not r["has_b_ultrasound"]]
    both = [r for r in all_rows if r["has_b_ultrasound"] and r["has_pathology"]]

    sampled_only_b = _pick_diverse(
        only_b,
        args.n_per_type,
        "b_image_count",
        args.seed,
        unique_counts=args.unique_counts,
        count_start=args.count_start,
    )
    sampled_only_pathology = _pick_diverse(
        only_pathology,
        args.n_per_type,
        "pathology_image_count",
        args.seed,
        unique_counts=args.unique_counts,
        count_start=args.count_start,
    )
    sampled_both = _pick_diverse(
        both,
        args.n_per_type,
        "total_image_count",
        args.seed,
        unique_counts=args.unique_counts,
        count_start=args.count_start,
    )

    columns = [
        "sickid",
        "b_image_count",
        "pathology_image_count",
        "total_image_count",
        "b_example_paths",
        "pathology_example_paths",
        "patient_dir",
    ]

    out_path = args.output.expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "only_b_ultrasound"
    ws2 = wb.create_sheet("only_pathology")
    ws3 = wb.create_sheet("both_modalities")

    def write_rows(ws, rows):
        ws.append(columns)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])

    write_rows(ws1, sampled_only_b)
    write_rows(ws2, sampled_only_pathology)
    write_rows(ws3, sampled_both)

    _fit_sheet_style(wb, Font)
    wb.save(out_path)

    copied_only_b = 0
    copied_only_pathology = 0
    copied_both = 0
    if args.export_sample_dir is not None:
        export_root = args.export_sample_dir.expanduser().resolve()
        export_root.mkdir(parents=True, exist_ok=True)
        copied_only_b = _export_sample_files(
            rows=sampled_only_b,
            sheet_name="only_b_ultrasound",
            count_key="b_image_count",
            export_root=export_root,
            include_b=True,
            include_pathology=False,
        )
        copied_only_pathology = _export_sample_files(
            rows=sampled_only_pathology,
            sheet_name="only_pathology",
            count_key="pathology_image_count",
            export_root=export_root,
            include_b=False,
            include_pathology=True,
        )
        copied_both = _export_sample_files(
            rows=sampled_both,
            sheet_name="both_modalities",
            count_key="total_image_count",
            export_root=export_root,
            include_b=True,
            include_pathology=True,
        )

    print(f"Done. xlsx saved to: {out_path}")
    print(
        "Category totals (all patients) => "
        f"only_b={len(only_b)}, only_pathology={len(only_pathology)}, both={len(both)}"
    )
    print(
        "Category sampled => "
        f"only_b={len(sampled_only_b)}, only_pathology={len(sampled_only_pathology)}, both={len(sampled_both)}"
    )
    if args.unique_counts:
        print("Sampling mode: unique-counts (one patient per distinct count value)")
    if args.export_sample_dir is not None:
        print(f"Done. sample files exported to: {args.export_sample_dir.expanduser().resolve()}")
        print(
            "Copied files => "
            f"only_b={copied_only_b}, only_pathology={copied_only_pathology}, both={copied_both}"
        )


if __name__ == "__main__":
    main()
